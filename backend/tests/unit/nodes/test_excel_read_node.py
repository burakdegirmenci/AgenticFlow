"""input.excel_read — spec-driven tests (docs/nodes/INPUT_EXCEL_READ_SPEC.md §10)."""

from __future__ import annotations

from pathlib import Path

import pytest

from app.engine.errors import NodeError
from app.nodes.input.excel_read import ExcelReadNode


@pytest.fixture
def node() -> ExcelReadNode:
    return ExcelReadNode()


# ---------------------------------------------------------------------------
# .xlsx fixtures
# ---------------------------------------------------------------------------
def _make_xlsx(path: Path, rows: list[list], sheet_name: str = "Sheet1") -> Path:
    """Write a tiny .xlsx for testing."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()
    return path


def _make_csv(path: Path, lines: list[str]) -> Path:
    path.write_text("\n".join(lines), encoding="utf-8-sig")
    return path


# ---------------------------------------------------------------------------
# Happy path — .xlsx with header
# ---------------------------------------------------------------------------
class TestXlsxWithHeader:
    async def test_reads_all_rows(self, node, execution_context, tmp_path) -> None:
        f = _make_xlsx(
            tmp_path / "test.xlsx",
            [["StokKodu", "Miktar"], ["SKU-001", 10], ["SKU-002", 20], ["SKU-003", 30]],
        )
        out = await node.execute(execution_context, {}, {"file_path": str(f), "header_row": True})
        assert out["count"] == 3
        assert out["items"][0] == {"StokKodu": "SKU-001", "Miktar": 10}
        assert out["items"][2] == {"StokKodu": "SKU-003", "Miktar": 30}
        assert out["columns"] == ["StokKodu", "Miktar"]

    async def test_column_map_renames_and_drops(self, node, execution_context, tmp_path) -> None:
        f = _make_xlsx(
            tmp_path / "test.xlsx",
            [["Stok Kodu", "Adet", "Fiyat"], ["A1", 5, 99.9], ["B2", 3, 149.0]],
        )
        out = await node.execute(
            execution_context,
            {},
            {
                "file_path": str(f),
                "column_map": {"SKU": "Stok Kodu", "Quantity": "Adet"},
            },
        )
        assert out["count"] == 2
        assert out["items"][0] == {"SKU": "A1", "Quantity": 5}
        assert "Fiyat" not in out["items"][0]  # dropped
        assert out["columns"] == ["SKU", "Quantity"]

    async def test_missing_source_column_gives_none(
        self, node, execution_context, tmp_path
    ) -> None:
        f = _make_xlsx(tmp_path / "test.xlsx", [["A", "B"], [1, 2]])
        out = await node.execute(
            execution_context,
            {},
            {"file_path": str(f), "column_map": {"x": "A", "y": "MISSING"}},
        )
        assert out["items"][0] == {"x": 1, "y": None}


# ---------------------------------------------------------------------------
# .xlsx without header (A, B, C naming)
# ---------------------------------------------------------------------------
class TestXlsxNoHeader:
    async def test_auto_column_names(self, node, execution_context, tmp_path) -> None:
        f = _make_xlsx(tmp_path / "test.xlsx", [[10, 20, 30], [40, 50, 60]])
        out = await node.execute(execution_context, {}, {"file_path": str(f), "header_row": False})
        assert out["count"] == 2
        assert out["items"][0] == {"A": 10, "B": 20, "C": 30}
        assert out["columns"] == ["A", "B", "C"]


# ---------------------------------------------------------------------------
# .csv
# ---------------------------------------------------------------------------
class TestCsv:
    async def test_reads_csv_with_header(self, node, execution_context, tmp_path) -> None:
        f = _make_csv(
            tmp_path / "test.csv",
            ["StokKodu,Miktar", "SKU-001,10", "SKU-002,20"],
        )
        out = await node.execute(execution_context, {}, {"file_path": str(f), "header_row": True})
        assert out["count"] == 2
        assert out["items"][0]["StokKodu"] == "SKU-001"
        assert out["items"][0]["Miktar"] == "10"  # CSV = all strings

    async def test_csv_without_header(self, node, execution_context, tmp_path) -> None:
        f = _make_csv(tmp_path / "test.csv", ["a,b", "c,d"])
        out = await node.execute(execution_context, {}, {"file_path": str(f), "header_row": False})
        assert out["items"][0] == {"A": "a", "B": "b"}
        assert out["items"][1] == {"A": "c", "B": "d"}


# ---------------------------------------------------------------------------
# Filtering
# ---------------------------------------------------------------------------
class TestFiltering:
    async def test_skip_empty_rows(self, node, execution_context, tmp_path) -> None:
        f = _make_xlsx(
            tmp_path / "test.xlsx",
            [["X"], ["val1"], [None], [""], ["val2"]],
        )
        out = await node.execute(
            execution_context,
            {},
            {"file_path": str(f), "skip_empty_rows": True},
        )
        assert out["count"] == 2
        assert [i["X"] for i in out["items"]] == ["val1", "val2"]

    async def test_max_rows_truncates(self, node, execution_context, tmp_path) -> None:
        f = _make_xlsx(
            tmp_path / "test.xlsx",
            [["V"]] + [[i] for i in range(100)],
        )
        out = await node.execute(
            execution_context,
            {},
            {"file_path": str(f), "max_rows": 5},
        )
        assert out["count"] == 5


# ---------------------------------------------------------------------------
# Error paths
# ---------------------------------------------------------------------------
class TestErrors:
    async def test_file_not_found(self, node, execution_context, tmp_path) -> None:
        # Absolute path that doesn't exist (uses tmp_path so it's inside a
        # known dir, bypassing the traversal check on the absolute-path branch).
        missing = tmp_path / "surely_missing.xlsx"
        with pytest.raises(NodeError, match="File not found"):
            await node.execute(execution_context, {}, {"file_path": str(missing)})

    async def test_unsupported_extension(self, node, execution_context, tmp_path) -> None:
        f = tmp_path / "old.xls"
        f.write_bytes(b"fake")
        with pytest.raises(NodeError, match="Unsupported format"):
            await node.execute(execution_context, {}, {"file_path": str(f)})

    async def test_empty_file_path(self, node, execution_context) -> None:
        with pytest.raises(NodeError, match="file_path is required"):
            await node.execute(execution_context, {}, {"file_path": ""})

    async def test_path_traversal_blocked(self, node, execution_context, tmp_path) -> None:
        with pytest.raises(NodeError, match="traversal"):
            await node.execute(execution_context, {}, {"file_path": "../../../etc/passwd"})


# ---------------------------------------------------------------------------
# Sheet selection
# ---------------------------------------------------------------------------
class TestSheetSelection:
    async def test_specific_sheet_name(self, node, execution_context, tmp_path) -> None:
        from openpyxl import Workbook

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Ignore"
        ws1.append(["wrong"])
        ws2 = wb.create_sheet("StokSayfasi")
        ws2.append(["SKU", "Qty"])
        ws2.append(["A1", 5])
        wb.save(tmp_path / "multi.xlsx")
        wb.close()

        out = await node.execute(
            execution_context,
            {},
            {"file_path": str(tmp_path / "multi.xlsx"), "sheet_name": "StokSayfasi"},
        )
        assert out["count"] == 1
        assert out["items"][0]["SKU"] == "A1"

    async def test_missing_sheet_name_raises(self, node, execution_context, tmp_path) -> None:
        f = _make_xlsx(tmp_path / "test.xlsx", [["a"], [1]])
        with pytest.raises(NodeError, match="Sheet.*not found"):
            await node.execute(
                execution_context,
                {},
                {"file_path": str(f), "sheet_name": "DoesNotExist"},
            )


# ---------------------------------------------------------------------------
# column_map as JSON string (UI round-trip)
# ---------------------------------------------------------------------------
class TestColumnMapJsonString:
    async def test_json_string_parsed(self, node, execution_context, tmp_path) -> None:
        f = _make_xlsx(tmp_path / "test.xlsx", [["A", "B"], [1, 2]])
        out = await node.execute(
            execution_context,
            {},
            {"file_path": str(f), "column_map": '{"x": "A"}'},
        )
        assert out["items"][0] == {"x": 1}
