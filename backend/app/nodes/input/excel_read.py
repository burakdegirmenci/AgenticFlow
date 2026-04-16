"""input.excel_read — Read rows from an Excel (.xlsx) or CSV file.

Spec: docs/nodes/INPUT_EXCEL_READ_SPEC.md

Turns a file on disk into the standard ``{"items": [...], "count": N}``
shape that every transform / ticimax / output node already consumes.
The user controls which columns to pick and how to rename them via
``column_map`` so downstream nodes get the field names they expect
without an extra ``transform.map`` step.

File resolution:
    ``file_path`` is resolved relative to ``<cwd>/uploads/``. Absolute paths
    are allowed but logged as a warning. Path traversal is blocked.

Supported formats:
    - ``.xlsx`` via openpyxl (``data_only=True`` — cached formula values).
    - ``.csv`` via stdlib csv.DictReader (UTF-8 with BOM tolerance).
    - ``.xls`` (legacy binary) is NOT supported.
"""

from __future__ import annotations

import csv
import json
import os
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from app.engine.context import ExecutionContext
from app.engine.errors import NodeError
from app.engine.node_base import BaseNode
from app.logging_config import get_logger
from app.nodes import register

logger = get_logger("agenticflow.nodes.input.excel_read")

_UPLOADS_DIR = Path(os.getcwd()) / "uploads"
_ALLOWED_EXTENSIONS = {".xlsx", ".csv"}


def _resolve_path(file_path: str) -> Path:
    """Resolve ``file_path`` to an absolute path inside the uploads dir.

    Raises ``NodeError`` on path traversal or missing file.
    """
    raw = file_path.strip()
    if not raw:
        raise NodeError("", "input.excel_read", "file_path is required")

    # Remember whether the user gave an absolute path (operator override)
    # vs a relative path (must stay inside uploads/).
    user_gave_absolute = Path(raw).is_absolute()

    candidate = Path(raw) if user_gave_absolute else (_UPLOADS_DIR / raw)
    resolved = candidate.resolve()

    # Security: block path traversal outside uploads/ and exports/.
    uploads_resolved = _UPLOADS_DIR.resolve()
    exports_resolved = (Path(os.getcwd()) / "exports").resolve()
    inside_safe_dir = str(resolved).startswith(str(uploads_resolved)) or str(resolved).startswith(
        str(exports_resolved)
    )

    if not inside_safe_dir:
        if not user_gave_absolute:
            # Relative path escaped uploads/ via ../.. → block.
            raise NodeError(
                "",
                "input.excel_read",
                f"Path traversal blocked: '{raw}' resolves outside uploads/",
            )
        # Absolute paths are allowed but warned (operator responsibility).
        logger.warning(
            "excel_read_absolute_path",
            extra={"resolved": str(resolved)},
        )

    if not resolved.exists():
        raise NodeError(
            "",
            "input.excel_read",
            f"File not found: {resolved}",
        )

    ext = resolved.suffix.lower()
    if ext not in _ALLOWED_EXTENSIONS:
        raise NodeError(
            "",
            "input.excel_read",
            f"Unsupported format '{ext}'. Use .xlsx or .csv.",
        )

    return resolved


def _coerce_cell(value: Any) -> Any:
    """Convert openpyxl / csv cell values to JSON-safe primitives."""
    if value is None:
        return None
    if isinstance(value, (int, float, bool)):
        return value
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    return str(value)


def _read_xlsx(
    path: Path,
    sheet_name: str,
    header_row: bool,
) -> tuple[list[str], list[list[Any]]]:
    """Read an .xlsx file. Returns (column_names, rows_as_lists)."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise NodeError(
                "",
                "input.excel_read",
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}",
            )
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows_raw: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        rows_raw.append([_coerce_cell(c) for c in row])

    wb.close()

    if not rows_raw:
        return [], []

    if header_row:
        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows_raw[0])]
        data = rows_raw[1:]
    else:
        ncols = max(len(r) for r in rows_raw) if rows_raw else 0
        headers = [chr(65 + i) if i < 26 else f"col_{i}" for i in range(ncols)]
        data = rows_raw

    return headers, data


def _read_csv(
    path: Path,
    header_row: bool,
) -> tuple[list[str], list[list[Any]]]:
    """Read a .csv file. Returns (column_names, rows_as_lists)."""
    # Detect BOM
    raw = path.read_bytes()
    text = raw.decode("utf-8-sig")

    lines = text.splitlines()
    if not lines:
        return [], []

    reader = csv.reader(lines)
    rows_raw = list(reader)

    if not rows_raw:
        return [], []

    if header_row:
        headers = [h.strip() or f"col_{i}" for i, h in enumerate(rows_raw[0])]
        data = rows_raw[1:]
    else:
        ncols = max(len(r) for r in rows_raw) if rows_raw else 0
        headers = [chr(65 + i) if i < 26 else f"col_{i}" for i in range(ncols)]
        data = rows_raw

    return headers, data


def _apply_column_map(
    headers: list[str],
    data: list[list[Any]],
    column_map: dict[str, str],
    skip_empty_rows: bool,
    max_rows: int,
) -> tuple[list[dict[str, Any]], list[str]]:
    """Apply column mapping + filtering. Returns (items, output_columns)."""
    if column_map:
        # Build source_index → output_name mapping.
        # column_map = {"OutputName": "SourceColumnName"}
        header_index = {h: i for i, h in enumerate(headers)}
        mappings: list[tuple[str, int | None]] = []
        for out_name, src_name in column_map.items():
            idx = header_index.get(src_name)
            mappings.append((out_name, idx))
        output_columns = [m[0] for m in mappings]
    else:
        # Pass through all columns.
        mappings = [(h, i) for i, h in enumerate(headers)]
        output_columns = list(headers)

    items: list[dict[str, Any]] = []
    for row in data:
        item: dict[str, Any] = {}
        for out_name, idx in mappings:
            if idx is not None and idx < len(row):
                item[out_name] = _coerce_cell(row[idx])
            else:
                item[out_name] = None

        if skip_empty_rows and all(v is None or v == "" for v in item.values()):
                continue

        items.append(item)
        if max_rows > 0 and len(items) >= max_rows:
            break

    return items, output_columns


@register
class ExcelReadNode(BaseNode):
    type_id = "input.excel_read"
    category = "input"
    display_name = "Excel / CSV Oku"
    description = (
        "Bir .xlsx veya .csv dosyasını okur ve satırları workflow'a "
        "aktarır. Kolon eşlemesi ile downstream node'ların beklediği "
        "alan adlarına dönüştürür."
    )
    icon = "file-spreadsheet"
    color = "#059669"

    input_schema = {"type": "object", "properties": {}}

    output_schema = {
        "type": "object",
        "properties": {
            "items": {"type": "array"},
            "count": {"type": "integer"},
            "source_file": {"type": "string"},
            "columns": {"type": "array"},
        },
    }

    config_schema = {
        "type": "object",
        "properties": {
            "file_path": {
                "type": "string",
                "title": "Dosya Yolu",
                "description": (
                    "uploads/ klasöründeki dosya adı veya tam yol. .xlsx ve .csv desteklenir."
                ),
                "default": "",
            },
            "sheet_name": {
                "type": "string",
                "title": "Sayfa Adı (Excel)",
                "description": "Boş = ilk sayfa. CSV için yok sayılır.",
                "default": "",
            },
            "header_row": {
                "type": "boolean",
                "title": "İlk Satır Başlık mı?",
                "description": (
                    "true = ilk satır kolon isimleri olarak kullanılır. "
                    "false = A, B, C... olarak adlandırılır."
                ),
                "default": True,
            },
            "column_map": {
                "type": "object",
                "title": "Kolon Eşlemesi",
                "description": (
                    "Çıkış alan adı → kaynak kolon adı. Örn: "
                    '{"StokKodu": "Stok Kodu", "Miktar": "Adet"}. '
                    "Boş = tüm kolonlar olduğu gibi geçer."
                ),
                "default": {},
            },
            "skip_empty_rows": {
                "type": "boolean",
                "title": "Boş Satırları Atla",
                "default": True,
            },
            "max_rows": {
                "type": "integer",
                "title": "Maksimum Satır",
                "description": "0 = sınırsız. Test için 10, production için 0.",
                "default": 0,
                "minimum": 0,
            },
        },
        "required": ["file_path"],
    }

    async def execute(
        self,
        context: ExecutionContext,
        inputs: dict[str, Any],
        config: dict[str, Any],
    ) -> dict[str, Any]:
        file_path = str(config.get("file_path", ""))
        sheet_name = str(config.get("sheet_name", "") or "")
        header_row = bool(config.get("header_row", True))
        skip_empty = bool(config.get("skip_empty_rows", True))
        max_rows = int(config.get("max_rows", 0))

        column_map_raw = config.get("column_map") or {}
        if isinstance(column_map_raw, str):
            try:
                column_map_raw = json.loads(column_map_raw)
            except (json.JSONDecodeError, TypeError):
                column_map_raw = {}
        column_map: dict[str, str] = {str(k): str(v) for k, v in column_map_raw.items()}

        resolved = _resolve_path(file_path)
        ext = resolved.suffix.lower()

        if ext == ".xlsx":
            headers, data = _read_xlsx(resolved, sheet_name, header_row)
        else:
            headers, data = _read_csv(resolved, header_row)

        items, output_columns = _apply_column_map(headers, data, column_map, skip_empty, max_rows)

        logger.info(
            "excel_read_complete",
            extra={
                "file": str(resolved),
                "format": ext,
                "raw_rows": len(data),
                "output_rows": len(items),
                "columns": output_columns,
            },
        )

        return {
            "items": items,
            "count": len(items),
            "source_file": str(resolved),
            "columns": output_columns,
        }
