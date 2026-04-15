"""Excel export node - writes array data to an .xlsx file via openpyxl.

Mirrors csv_export's discovery rules:
    * If `source_field` is set, walk the dotted path (e.g. `result.UrunList`).
    * Otherwise recursively search parent outputs for the first list-of-dicts.

Output lands in `<backend cwd>/exports/<filename>_<YYYYMMDD_HHMMSS>.xlsx`.
"""

import os
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.engine.context import ExecutionContext
from app.engine.node_base import BaseNode
from app.nodes import register


def _walk_dotted(obj: Any, path: str) -> Any:
    cur = obj
    for part in path.split("."):
        if part == "":
            continue
        if isinstance(cur, dict):
            cur = cur.get(part)
        elif isinstance(cur, list):
            try:
                idx = int(part)
                cur = cur[idx]
            except (ValueError, IndexError):
                return None
        else:
            return None
        if cur is None:
            return None
    return cur


def _find_first_list_of_dicts(obj: Any, max_depth: int = 6) -> list[dict] | None:
    if max_depth < 0:
        return None
    if isinstance(obj, list):
        if obj and isinstance(obj[0], dict):
            return obj  # type: ignore[return-value]
        return None
    if isinstance(obj, dict):
        for v in obj.values():
            if isinstance(v, list) and v and isinstance(v[0], dict):
                return v  # type: ignore[return-value]
        for v in obj.values():
            found = _find_first_list_of_dicts(v, max_depth - 1)
            if found is not None:
                return found
    return None


def _coerce_cell(value: Any) -> Any:
    """openpyxl accepts str/int/float/bool/None/datetime natively.

    Everything else (dict, list, zeep objects) is stringified.
    """
    if value is None or isinstance(value, (str, int, float, bool, datetime, date)):
        return value
    return str(value)


@register
class ExcelExportNode(BaseNode):
    type_id = "output.excel_export"
    category = "output"
    display_name = "Excel Export"
    description = (
        "Parent node çıktısındaki listeyi .xlsx dosyasına yazar. "
        "Dosya `backend/exports/<ad>_<zaman>.xlsx` yoluna kaydedilir."
    )
    icon = "file-spreadsheet"
    color = "#16a34a"

    config_schema = {
        "type": "object",
        "properties": {
            "filename": {
                "type": "string",
                "title": "Dosya Adı (uzantısız)",
                "default": "export",
            },
            "sheet_name": {
                "type": "string",
                "title": "Sayfa Adı",
                "default": "Sheet1",
            },
            "source_field": {
                "type": "string",
                "title": "Kaynak Alan (dotted path)",
                "description": (
                    "Örn: 'result.UrunList' veya 'siparisler'. "
                    "Boşsa otomatik olarak ilk dict-listesi bulunur."
                ),
                "default": "",
            },
            "freeze_header": {
                "type": "boolean",
                "title": "Başlık Satırını Sabitle",
                "default": True,
            },
        },
    }

    output_schema = {
        "type": "object",
        "properties": {
            "file_path": {"type": "string"},
            "rows_written": {"type": "integer"},
        },
    }

    async def execute(
        self,
        context: ExecutionContext,
        inputs: dict[str, Any],
        config: dict[str, Any],
    ) -> dict[str, Any]:
        source_field = (config.get("source_field") or "").strip()
        rows: list[dict] = []
        source_parent: str = ""

        for parent_id, parent_output in inputs.items():
            if not isinstance(parent_output, (dict, list)):
                continue
            if source_field:
                candidate = _walk_dotted(parent_output, source_field)
                if isinstance(candidate, list) and candidate:
                    rows = candidate
                    source_parent = parent_id
                    break
            else:
                found = _find_first_list_of_dicts(parent_output)
                if found:
                    rows = found
                    source_parent = parent_id
                    break

        if not rows:
            return {
                "file_path": "",
                "rows_written": 0,
                "note": (
                    f"no list-of-dicts found in inputs from {list(inputs.keys())}"
                    if not source_field
                    else f"source_field '{source_field}' did not resolve to a list"
                ),
            }

        exports_dir = os.path.join(os.getcwd(), "exports")
        os.makedirs(exports_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = (config.get("filename") or "export") or "export"
        file_path = os.path.join(exports_dir, f"{filename}_{ts}.xlsx")

        # Collect column keys (union, preserving first-seen order)
        keys: list[str] = []
        seen: set[str] = set()
        for row in rows:
            if not isinstance(row, dict):
                continue
            for k in row.keys():
                if k not in seen:
                    seen.add(k)
                    keys.append(k)

        wb = Workbook()
        ws = wb.active
        ws.title = (config.get("sheet_name") or "Sheet1")[:31]  # Excel sheet name cap

        # Header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
        for col_idx, key in enumerate(keys, start=1):
            cell = ws.cell(row=1, column=col_idx, value=key)
            cell.font = header_font
            cell.fill = header_fill

        # Data rows
        for row_idx, row in enumerate(rows, start=2):
            if not isinstance(row, dict):
                continue
            for col_idx, key in enumerate(keys, start=1):
                ws.cell(row=row_idx, column=col_idx, value=_coerce_cell(row.get(key)))

        # Column width auto-fit (capped)
        for col_idx, key in enumerate(keys, start=1):
            max_len = len(str(key))
            for row in rows[:200]:  # sample for performance
                if isinstance(row, dict):
                    v = row.get(key)
                    if v is not None:
                        max_len = max(max_len, min(len(str(v)), 60))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

        if bool(config.get("freeze_header", True)):
            ws.freeze_panes = "A2"

        wb.save(file_path)

        return {
            "file_path": file_path,
            "rows_written": len(rows),
            "source_parent": source_parent,
        }
