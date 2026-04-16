"""File upload + preview endpoints for runtime workflow inputs.

Spec: docs/RUNTIME_INPUT_SPEC.md §2.2, docs/EXCEL_PREVIEW_MAPPING_SPEC.md

Upload: saves file to ``<cwd>/uploads/`` with timestamp prefix.
Preview: reads first N rows so the UI can show a column-mapping dialog.
"""

from __future__ import annotations

import csv
import os
from datetime import datetime
from pathlib import Path
from typing import Any

from fastapi import APIRouter, HTTPException, UploadFile

from app.logging_config import get_logger

logger = get_logger("agenticflow.uploads")

router = APIRouter()

_UPLOADS_DIR = Path(os.getcwd()) / "uploads"
_ALLOWED_EXTENSIONS = {".xlsx", ".csv"}
_MAX_SIZE_BYTES = 10 * 1024 * 1024  # 10 MB


@router.post("")
async def upload_file(file: UploadFile) -> dict:
    """Upload a file for use in workflow runs.

    Returns ``{"filename": "<timestamped>", "original": "<original>", "size": N}``.
    The ``filename`` value is what goes into ``input_data.file`` when
    triggering a workflow that uses ``{{trigger_input.file}}``.
    """
    if not file.filename:
        raise HTTPException(400, "No filename provided")

    ext = Path(file.filename).suffix.lower()
    if ext not in _ALLOWED_EXTENSIONS:
        raise HTTPException(
            400,
            f"Unsupported file type '{ext}'. Allowed: {', '.join(sorted(_ALLOWED_EXTENSIONS))}",
        )

    # Read + enforce size limit
    content = await file.read()
    if len(content) > _MAX_SIZE_BYTES:
        raise HTTPException(
            413,
            f"File too large ({len(content):,} bytes). Max: {_MAX_SIZE_BYTES:,} bytes.",
        )

    # Timestamped filename to avoid collisions
    stem = Path(file.filename).stem
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = f"{stem}_{ts}{ext}"

    _UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    dest = _UPLOADS_DIR / safe_name
    dest.write_bytes(content)

    logger.info(
        "file_uploaded",
        extra={
            "original": file.filename,
            "saved_as": safe_name,
            "size": len(content),
        },
    )

    return {
        "filename": safe_name,
        "original": file.filename,
        "size": len(content),
    }


def _coerce_cell(value: Any) -> Any:
    """Convert cell to JSON-safe primitive."""
    if value is None:
        return None
    if isinstance(value, (int, float, bool)):
        return value
    from datetime import date, datetime as dt, time

    if isinstance(value, dt):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    return str(value)


@router.get("/{filename}/preview")
def preview_file(filename: str, rows: int = 5) -> dict:
    """Preview the first N rows of an uploaded file.

    Returns column names + sample data so the UI can render a
    column-mapping dialog before the workflow runs.
    """
    path = (_UPLOADS_DIR / filename).resolve()
    if not str(path).startswith(str(_UPLOADS_DIR.resolve())):
        raise HTTPException(400, "Invalid filename")
    if not path.exists():
        raise HTTPException(404, f"File not found: {filename}")

    ext = path.suffix.lower()
    if ext not in _ALLOWED_EXTENSIONS:
        raise HTTPException(400, f"Unsupported: {ext}")

    columns: list[str] = []
    sample_rows: list[dict[str, Any]] = []
    total_rows = 0

    if ext == ".xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        all_rows: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append([_coerce_cell(c) for c in row])
        wb.close()

        if all_rows:
            columns = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(all_rows[0])]
            data = all_rows[1:]
            total_rows = len(data)
            for row in data[: min(rows, 10)]:
                item = {}
                for i, col in enumerate(columns):
                    item[col] = row[i] if i < len(row) else None
                sample_rows.append(item)

    elif ext == ".csv":
        raw = path.read_bytes().decode("utf-8-sig")
        lines = raw.splitlines()
        if lines:
            reader = csv.reader(lines)
            all_csv = list(reader)
            if all_csv:
                columns = [h.strip() or f"col_{i}" for i, h in enumerate(all_csv[0])]
                data_csv = all_csv[1:]
                total_rows = len(data_csv)
                for row in data_csv[: min(rows, 10)]:
                    item = {}
                    for i, col in enumerate(columns):
                        item[col] = row[i] if i < len(row) else None
                    sample_rows.append(item)

    return {
        "filename": filename,
        "columns": columns,
        "sample_rows": sample_rows,
        "total_rows": total_rows,
    }
